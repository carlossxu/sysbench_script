import re
import subprocess
import os
import sys
from openpyxl import Workbook, load_workbook

def cpu_test():
    if os.path.exists("cpu_run_results") is False:
        subprocess.run("mkdir cpu_run_results", capture_output=True, text=True, shell=True)
    for i in range(20):

        tmp_file_name = "cpu_run_tmp{}".format(i+1)

        subprocess.run("sysbench cpu --threads=1 --time=60 run > ./cpu_run_results/{}".format(tmp_file_name), capture_output=True, text=True, shell=True)

        tmp_file_path = "./cpu_run_results/{}".format(tmp_file_name)
        with open(tmp_file_path, "r") as file:
            data = file.read()

        # pattern = r"events/s (eps):\s*:\s*([\d.]+)"
        pattern = r"events/s \(eps\):\s+(\d+\.\d+)"
        result_iterations_per_sec = re.search(pattern, data)

        if result_iterations_per_sec:

            print("Test no.{} has already done. The result is {} eps.".format(i+1, float(result_iterations_per_sec.group(1))))

            try:
                wb = load_workbook("cpu_test_result.xlsx")
            except FileNotFoundError:
                wb = Workbook()

            sheet = wb.active
            sheet["A1"] = "events/sec"
            sheet["B1"] = "thread_number"
            sheet["C1"] = "architecture"
            next_row = sheet.max_row + 1
            sheet[f"A{next_row}"] = float(result_iterations_per_sec.group(1))
            sheet[f"B{next_row}"] = 1

            arc = subprocess.check_output("uname -m", shell=True)
            sheet[f"C{next_row}"] = arc.decode('utf-8').strip()
            wb.save("cpu_test_result.xlsx")

        # if os.path.exists(tmp_file_path):
        #     os.remove(tmp_file_path)

# sysbench memtest的一个event是读写一个内存块
def mem_test():
    if os.path.exists("mem_run_results") is False:
        subprocess.run("mkdir mem_run_results", capture_output=True, text=True, shell=True)
    for i in range(20):
        # 内存读性能和写性能都测一遍
        if i < 10:
            oper = "read"
        else:
            oper = "write"

        tmp_file_name = "mem_{}_run_tmp{}".format(oper, i+1)
        subprocess.run("sysbench memory --threads=1 --memory-oper=read --time=60 run > ./mem_run_results/{}".format(tmp_file_name), capture_output=True, text=True, shell=True)
        
        tmp_file_path = "./mem_run_results/{}".format(tmp_file_name)
        with open(tmp_file_path, "r") as file:
            data = file.read()

        pattern1 = r'(\d+\.\d{2}) MiB/sec'
        pattern2 = r"events/s \(eps\):\s+(\d+\.\d+)"
        result_MiB_per_sec = re.search(pattern1, data)
        result_iterations_per_sec = re.search(pattern2, data)

        if result_MiB_per_sec and result_iterations_per_sec:

            print("Test no.{} has already done. The {}_result is {} MiB/sec.".format(i+1, oper, float(result_MiB_per_sec.group(1))))

            try:
                wb = load_workbook("mem_test_result.xlsx")
            except FileNotFoundError:
                wb = Workbook()

            sheet = wb.active
            sheet["A1"] = "MiB/sec"
            sheet["B1"] = "events/sec"
            sheet["C1"] = "thread_number"
            sheet["D1"] = "operation"
            sheet["E1"] = "architecture"
            next_row = sheet.max_row + 1
            sheet[f"A{next_row}"] = float(result_MiB_per_sec.group(1))
            sheet[f"B{next_row}"] = float(result_iterations_per_sec.group(1))
            sheet[f"C{next_row}"] = 1
            sheet[f"D{next_row}"] = oper
            arc = subprocess.check_output("uname -m", shell=True)
            sheet[f"E{next_row}"] = arc.decode('utf-8').strip()

            wb.save("mem_test_result.xlsx")

# 提取匹配到的值
def extract_values(match):
    if match:
        iops = float(match.group(1))
        mbps = float(match.group(2))
        return iops, mbps
    else:
        return None, None
    
def fileio_seqrw_test():
    if os.path.exists("fileio_run_results") is False:
        subprocess.run("mkdir fileio_run_results", capture_output=True, text=True, shell=True)
    for i in range(10):
        read_iops_result, write_iops_result, fsync_iops_result = 0, 0, 0
        read_mbps_result, write_mbps_result = 0, 0
        for j in range(2):
            if j == 0:
                oper = "seqwr"
            else:
                oper = "seqrd"
            tmp_file_name = "mem_{}_run_tmp{}".format(oper, i+1)
            subprocess.run("sysbench memory --threads=1 --memory-oper=read --time=60 run > ./mem_run_results/{}".format(tmp_file_name), capture_output=True, text=True, shell=True)
            
            tmp_file_path = "./mem_run_results/{}".format(tmp_file_name)
            with open(tmp_file_path, "r") as file:
                data = file.read()

            read_match = re.search(r'read:\s+IOPS=([\d.]+)\s+([\d.]+)\s+MiB/s', data)
            write_match = re.search(r'write:\s+IOPS=([\d.]+)\s+([\d.]+)\s+MiB/s', data)
            fsync_match = re.search(r'fsync:\s+IOPS=([\d.]+)', data)

            read_iops, read_mbps = extract_values(read_match)
            write_iops, write_mbps = extract_values(write_match)
            fsync_iops = float(fsync_match.group(1))

            read_iops_result = max(read_iops_result, read_iops)
            write_iops_result = max(write_iops_result, write_iops)
            fsync_iops_result = max(fsync_iops_result, fsync_iops)
            read_mbps_result = max(read_mbps_result, read_mbps)
            write_mbps_result = max(write_mbps_result, write_mbps)

        try:
            wb = load_workbook("filio_test_result.xlsx")
        except FileNotFoundError:
            wb = Workbook()

        sheet = wb.active
        sheet["A1"] = "read_iops"
        sheet["B1"] = "read_MiBps"
        sheet["C1"] = "write_iops"
        sheet["D1"] = "write_MiBps"
        sheet["E1"] = "fsync_iops"
        sheet["F1"] = "operation"
        sheet["G1"] = "architecture"

        next_row = sheet.max_row + 1
        sheet[f"A{next_row}"] = read_iops_result
        sheet[f"B{next_row}"] = read_mbps_result
        sheet[f"C{next_row}"] = write_iops_result
        sheet[f"D{next_row}"] = write_mbps_result
        sheet[f"E{next_row}"] = fsync_iops_result
        sheet[f"F{next_row}"] = oper
        arc = subprocess.check_output("uname -m", shell=True)
        sheet[f"G{next_row}"] = arc.decode('utf-8').strip()

        wb.save("fileio_test_result.xlsx")

        print("Test no.{} has already done..".format(i+1))



def fileio_test():
    if os.path.exists("fileio_run_results") is False:
        subprocess.run("mkdir fileio_run_results", capture_output=True, text=True, shell=True)
    for i in range(2):
        # 内存读性能和写性能都测一遍
        if i == 0:
            for j in range(10)
        else:


        tmp_file_name = "mem_{}_run_tmp{}".format(oper, i+1)
        subprocess.run("sysbench memory --threads=1 --memory-oper=read --time=60 run > ./mem_run_results/{}".format(tmp_file_name), capture_output=True, text=True, shell=True)
        
        tmp_file_path = "./mem_run_results/{}".format(tmp_file_name)
        with open(tmp_file_path, "r") as file:
            data = file.read()

        pattern1 = r'(\d+\.\d{2}) MiB/sec'
        pattern2 = r"events/s \(eps\):\s+(\d+\.\d+)"
        result_MiB_per_sec = re.search(pattern1, data)
        result_iterations_per_sec = re.search(pattern2, data)

        if result_MiB_per_sec and result_iterations_per_sec:

            print("Test no.{} has already done. The {}_result is {} MiB/sec.".format(i+1, oper, float(result_MiB_per_sec.group(1))))

            try:
                wb = load_workbook("mem_test_result.xlsx")
            except FileNotFoundError:
                wb = Workbook()

            sheet = wb.active
            sheet["A1"] = "MiB/sec"
            sheet["B1"] = "events/sec"
            sheet["C1"] = "thread_number"
            sheet["D1"] = "operation"
            sheet["E1"] = "architecture"
            next_row = sheet.max_row + 1
            sheet[f"A{next_row}"] = float(result_MiB_per_sec.group(1))
            sheet[f"B{next_row}"] = float(result_iterations_per_sec.group(1))
            sheet[f"C{next_row}"] = 1
            sheet[f"D{next_row}"] = oper
            arc = subprocess.check_output("uname -m", shell=True)
            sheet[f"E{next_row}"] = arc.decode('utf-8').strip()

            wb.save("mem_test_result.xlsx")


def main():
    args = sys.argv
    if len(args) != 2:
        print("Usage: python3 benchmark.py cpu/mem/fileio.")
        return
    test_dict = {
        "cpu": cpu_test,
        "mem": mem_test,
        "fileio": fileio_test
    }
    
    test_func = test_dict.get(args[1], None)
    if test_func is not None:
        test_func()
    else:
        print("Invalid test.")

if __name__ == "__main__":
    main()
    

    




